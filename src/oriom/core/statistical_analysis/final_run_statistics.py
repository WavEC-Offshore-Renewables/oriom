import os 
import pandas as pd
import openpyxl
import logging
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
from copy import deepcopy


from oriom.utils import aux_functions
from oriom.core.functions.graphs import final_economic_graphs
from oriom.core.functions.graphs import report_graphs
try:
    from oriom.core.functions.private.KPI_Insight import KPI_Insight
except ImportError:
    KPI_Insight = None


def return_statistics_runs(
        n_lifetime: int,
        find_element_class: dict,
        results_dict: object,
        fuel_add: dict,
        mobilisation_add: dict,
        electricity_cost_dict: dict,
        n_runs: int,
        vessels: list,
        operations_total: list,
        recycled: bool = False,
        save_dir: str = None
    ):

    """ 
    Create averaged data for each simulation. Consider yearly costs, total costs and ctv costs strategy.

    Save all resutls in sheets of excels file and create graphs
    
    Args:
        n_lifetime (int): number of year of lifetime,
        find_element_class (Find_Element): Initialized instance that provides fast access to operations, vessels and failures via internal dictionaries.
        results_dict (object): Object of class `Results`
        fuel_add (dict): dictionary of vessel_type and fuel yearly cost to add 
        mobilisation_add (dict): dictionary of vessel_type and mobilisation yearly cost to add 
        electricity_cost_dict: (dict): jey of tech and value cost of electricity per kWh
        n_runs (int): number of simulations
        vessels (:obj: `list`): List of object with attribute `id` for class `Vessels`
        operations_total (:obj: `list`): List of object for all the classes `Operations`
        save_dir (str, optional): directory to averaged results. Defaults to None.
    """
    
    
    def restructure_df_year(kpi_om_year_final: pd.DataFrame) -> pd.DataFrame:
        """
        Create a MultiIndex for yearly KPIs from columns like '1990', '1990.1', etc.
        """
        def parse_column(col):
            if '.' in col:
                year, suffix = col.split('.')
                metric = 'n_days'
            else:
                year = col
                metric = 'direct_costs'
            return (int(year), metric)
        
        vessel_ids = kpi_om_year_final['vessel_id']
        data = kpi_om_year_final.drop(columns='vessel_id')
        data = data.iloc[1:]
        data.reset_index(drop=True)

        multi_cols = pd.MultiIndex.from_tuples(
            [parse_column(col) for col in data.columns],
            names=['year', 'metric']
        )

        data.columns = multi_cols
        data = data.sort_index(axis=1, level=0)
        data.insert(0, 'vessel_id', vessel_ids)

        return data

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Lifetime_results"

    lifetime_fixed_port_cost, lifetime_technician_cost , lifetime_vessels_cost, lifetime_rov_cost  = [], [], [], []
    lifetime_repair_cost, lifetime_fixed_tech_cost,  lifetime_mobilisation_cost, lifetime_fixed_insurance_cost  = [], [], [], []
    final_cost, final_energy = {}, {}

    # Averaged total costs
    for kpi_tot_sim in results_dict.dfs_tot_cost_list:
        kpi_tot = kpi_tot_sim.iloc[-1]
        kpi_tot_sim = kpi_tot_sim.iloc[:-1]
        kpi_tot_sim['lifetime_direct_costs'] = kpi_tot_sim['lifetime_direct_costs'].astype(float)
        fuel_to_add = sum(fuel_add.values())
        mobilisation_to_add = sum(mobilisation_add.values())
        
        lifetime_fixed_insurance_cost.append(kpi_tot_sim.loc[kpi_tot_sim['vessel_id'] == 'insurance', 'lifetime_direct_costs'].values[0])
        lifetime_fixed_port_cost.append(kpi_tot_sim.loc[kpi_tot_sim['vessel_id'] == 'port', 'lifetime_direct_costs'].values[0])
        lifetime_fixed_tech_cost.append(kpi_tot_sim.loc[kpi_tot_sim['vessel_id'] == 'technician', 'lifetime_direct_costs'].values[0])
        lifetime_technician_cost.append(kpi_tot['tot_technicians_costs'])
        lifetime_vessels_cost.append(kpi_tot['tot_vessel_costs'] + n_lifetime * fuel_to_add)
        lifetime_mobilisation_cost.append(kpi_tot['tot_mobilization_costs'] + n_lifetime * mobilisation_to_add)
        lifetime_rov_cost.append(kpi_tot['tot_rov_costs'])
        lifetime_repair_cost.append(kpi_tot['tot_part_costs'])


    lifetime_fixed_insurance_cost_avg = sum(lifetime_fixed_insurance_cost)/len(lifetime_fixed_insurance_cost)
    lifetime_technician_cost_avg = sum(lifetime_technician_cost)/len(lifetime_technician_cost)
    lifetime_vessels_cost_avg = sum(lifetime_vessels_cost)/len(lifetime_vessels_cost)
    lifetime_mobilisation_cost_avg = sum(lifetime_mobilisation_cost)/len(lifetime_mobilisation_cost)
    lifetime_rov_cost_avg = sum(lifetime_rov_cost)/len(lifetime_rov_cost)
    lifetime_repair_cost_avg = sum(lifetime_repair_cost)/len(lifetime_repair_cost)
    lifetime_fixed_tech_cost_avg = sum(lifetime_fixed_tech_cost)/len(lifetime_fixed_tech_cost)
    lifetime_fixed_port_cost_avg = sum(lifetime_fixed_port_cost)/len(lifetime_fixed_port_cost)


    dict_cost = {
        'lifetime_vessels_cost €':lifetime_vessels_cost_avg,
        'lifetime_mobilisation_cost €':lifetime_mobilisation_cost_avg,
        'lifetime_rov_cost €':lifetime_rov_cost_avg,
        'lifetime_technician_cost €':lifetime_technician_cost_avg,
        'lifetime_repair_cost €':lifetime_repair_cost_avg,
        'lifetime_fixed_port_cost €':lifetime_fixed_port_cost_avg,
        'lifetime_fixed_tech_cost €':lifetime_fixed_tech_cost_avg,
        'lifetime_fixed_insurance_cost €':lifetime_fixed_insurance_cost_avg,
    }
    
    df_cost = pd.DataFrame([{'cost_type': k, 'value': v} for k, v in dict_cost.items()])
    final_cost['lifetime_direct_cost'] = df_cost['value'].sum()

    # Create graphs
    final_economic_graphs.lifetime_cost(df = df_cost, title = 'Lifetime costs', save_dir = save_dir, file_name='lifetime cost')
    df_cost_yearly = deepcopy(df_cost)
    final_cost['yearly_direct_cost'] = final_cost['lifetime_direct_cost']/n_lifetime
    df_cost_yearly['value'] = df_cost_yearly['value'] / n_lifetime
    final_economic_graphs.lifetime_cost(df = df_cost_yearly, title = 'Yearly costs', save_dir = save_dir, file_name='yearly cost')

    # Save on excel
    df_cost.loc[len(df_cost)] = ['lifetime_direct_cost', df_cost['value'].sum()]
    
    ws2 = wb.create_sheet(title="Lifetime_costs")
    for row in dataframe_to_rows(df_cost, index=False, header=True):
        ws2.append(row)


    # Averaged yearly values
    # If read from excel file need to create multiindex
    if recycled:
        for i in range(len(results_dict.dfs_tot_yearly_cost_list)):
            results_dict.dfs_tot_yearly_cost_list[i] = restructure_df_year(results_dict.dfs_tot_yearly_cost_list[i])

    array_stack = np.stack([df.iloc[:, 1:].to_numpy(dtype=np.float64) for df in results_dict.dfs_tot_yearly_cost_list], axis=0)

    # Evaluate the average for each cell
    mean_array = np.nanmean(array_stack, axis=0)

    # Reconstruct the df
    vessel_ids = results_dict.dfs_tot_yearly_cost_list[0].iloc[:, 0].values  # first column "vessel_id"
    column_names = results_dict.dfs_tot_yearly_cost_list[0].columns[1:]  # exclude 'vessel_id'
    df_cost_yearly = pd.DataFrame(mean_array, columns=column_names)
    df_cost_yearly.insert(0, "vessel_id", vessel_ids)

    years = sorted(set(col[0] for col in df_cost_yearly.columns if col[1] == 'direct_costs'))

    for ves in fuel_add.keys():
        for year in years:
            col_key = (year, 'direct_costs')
            if col_key in df_cost_yearly.columns:
                mask = df_cost_yearly['vessel_id'] == ves
                df_cost_yearly.loc[mask, col_key] += fuel_add[ves]
    for ves in mobilisation_add.keys():
        for year in years:
            col_key = (year, 'direct_costs')
            if col_key in df_cost_yearly.columns:
                mask = df_cost_yearly['vessel_id'] == ves
                df_cost_yearly.loc[mask, col_key] += mobilisation_add[ves]

    ws3 = wb.create_sheet(title="Yearly_costs")
    for row in dataframe_to_rows(df_cost_yearly, index=False, header=True):
        ws3.append(row)
    final_economic_graphs.yearly_vessel_cost(df = df_cost_yearly, title = 'Yearly costs', find_element_class = find_element_class, save_dir = save_dir)

    # Average cost ctv
    if results_dict.dfs_ctv_list and any(not df.empty for df in results_dict.dfs_ctv_list):
        df_ctv = sum(results_dict.dfs_ctv_list) / len(results_dict.dfs_ctv_list)
        df_ctv['n_vessel'] = range(len(df_ctv))
        df_ctv = df_ctv[['n_vessel'] + [col for col in df_ctv.columns if col != 'n_vessel']]
        ws4 = wb.create_sheet(title="CTV_yearly_strategy")
        for row in dataframe_to_rows(df_ctv, index=False, header=True):
            ws4.append(row)
        final_economic_graphs.plot_ctv_annual_chart_strategy(df_ctv = df_ctv, save_dir = save_dir)

    # Average type cost
    if results_dict.kpi_om_type_cost_list:
        kpi_om_type_cost_concat = pd.concat(results_dict.kpi_om_type_cost_list)
        kpi_om_type_avg = kpi_om_type_cost_concat.groupby('description', as_index=False)['values'].mean()
        final_economic_graphs.pie_chart(
            df = kpi_om_type_avg,
            description = 'description',
            value = 'values',
            title = 'Lifetime Inspections vs Corrections costs',
            tot_in_title = True,
            title_unit = 'M€',
            coefficient_number = 1e6,
            legend_title = 'Operation type',
            legend_bbox_to_anchor = (1.05, 0.9),
            text = 'mobilisation cost not included',
            save_dir = save_dir,
            file_name_save = 'O&M type cost' 
        )

    # Availability year average and year_month avg
    final_cost['AEP_kWh'] = 0
    final_cost['En_loss_€'] = 0
    final_cost['En_loss_kWh'] = 0
    
    for k in results_dict.dfs_energy_yearly_dict.keys():
        if results_dict.dfs_energy_yearly_dict[k]:
            df_yearly_mean = (
                pd.concat([
                        df.astype(float) for df in results_dict.dfs_energy_yearly_dict[k]],
                        axis=0, 
                        keys=range(len(results_dict.dfs_energy_yearly_dict[k]
                    )))
                .groupby(level=1)
                .mean()
                .reset_index(drop=True)
            )
            aux_functions.save_file_csv(df_to_save = df_yearly_mean, save_dir = save_dir, filename = f'{k}.csv', indexing = True)
            report_graphs.farm_availability(df = df_yearly_mean, name_file = k[17:], save_dir = save_dir)
            if 'pv' in k:
                electricity_cost_tech = electricity_cost_dict['pv']
                tech = 'pv'
            elif 'wind' in k:
                electricity_cost_tech = electricity_cost_dict['wt']
                tech = 'wind'
            elif 'wave' in k:
                electricity_cost_tech = electricity_cost_dict['wec']
                tech = 'wave'

            final_energy[f'{tech}_En_max_kWh'] = df_yearly_mean['En_max_kWh'].sum()
            final_energy[f'{tech}_En_loss_kWh'] = df_yearly_mean['En_loss_kWh'].sum()
            final_energy[f'{tech}_En_loss_€'] = final_energy[f'{tech}_En_loss_kWh'] * electricity_cost_tech/1e3  #eur/MWh
            final_energy[f'{tech}_En_produced_kWh'] = final_energy[f'{tech}_En_max_kWh'] - final_energy[f'{tech}_En_loss_kWh']
            final_energy[f'{tech}_AEP_kWh'] = final_energy[f'{tech}_En_produced_kWh']/n_lifetime
            final_energy[f'{tech}_En_availability_%'] = df_yearly_mean['En_availability'].mean()
            final_energy[f'{tech}_Time_availability_%'] = df_yearly_mean['Time_availability'].mean()
            final_cost['AEP_kWh'] += final_energy[f'{tech}_AEP_kWh']
            final_cost['En_loss_€'] += final_energy[f'{tech}_En_loss_€']
            final_cost['En_loss_kWh'] += final_energy[f'{tech}_En_loss_kWh']
            report_graphs.indirect_costs_per_year(df = df_yearly_mean, electricity_price = electricity_cost_tech, name_file = k[17:], save_dir = save_dir)
    
    # Monthly
    combined = {}
    for k in results_dict.dfs_energy_yearly_month_dict.keys():
        if results_dict.dfs_energy_yearly_month_dict[k]:
            df_monthly_mean = (
                pd.concat(
                        [df.astype(float) for df in results_dict.dfs_energy_yearly_month_dict[k]], 
                        axis=0, 
                        keys=range(len(results_dict.dfs_energy_yearly_month_dict[k]
                    )))
                .groupby(level=1)
                .mean()
                .reset_index(drop=True)
            )
            combined.update({k:df_monthly_mean})

            aux_functions.save_file_csv(df_to_save = df_monthly_mean, save_dir = save_dir, filename = f'{k}.csv')
            report_graphs.energy_yield(df = df_monthly_mean, name_file = k[18:], save_dir = save_dir)

    if len(combined.keys())>1:
        report_graphs.energy_yield_combined(dfs = combined, save_dir = save_dir)
    
    if KPI_Insight is not None:
        # Add insights from log events
        run_insight = KPI_Insight(N_SIMULATION = n_runs, n_lifetime = n_lifetime)
        cost_insight, vessel_insight = run_insight.kpi_insight(results_dict = results_dict, vessels = vessels, operations_total = operations_total)   
        
        if not vessel_insight.empty:
            final_cost['reuse_ctv'] = vessel_insight.loc["ctv", "reuse %"]
            final_cost['merge_ctv'] = vessel_insight.loc["ctv", "merge %"]
            final_cost['ctv_effective_day_y'] = vessel_insight.loc["ctv", "yearly day effective"]
    else:
        cost_insight = pd.DataFrame()
        vessel_insight = pd.DataFrame()

    df_lifetime_cost = pd.DataFrame({
        'Descriptions': final_cost.keys(),
        'Values': final_cost.values(),
    })

    for row in dataframe_to_rows(df_lifetime_cost, index=False, header=True):
        ws.append(row)

    if final_energy:
        df_energy_production = pd.DataFrame({
                'Descriptions': final_energy.keys(),
                'Values': final_energy.values(),
            })
        ws5 = wb.create_sheet(title="Energy_results")
        for row in dataframe_to_rows(df_energy_production, index=False, header=True):
            ws5.append(row)
    
        wb._sheets = [ws5] + [s for s in wb._sheets if s != ws5]

    # Reorder sheets and create final file
    wb._sheets = [ws] + [s for s in wb._sheets if s != ws]
    wb.save(os.path.join(save_dir,"Average_results.xlsx"))
    
    if (KPI_Insight is not None and 
        (vessel_insight is not None and not vessel_insight.empty)
         or (cost_insight is not None and not cost_insight.empty)
    ):
        with pd.ExcelWriter(
            os.path.join(save_dir, 'KPI_insight.xlsx'), 
            engine='openpyxl'
        ) as writer:
            v_ins, c_ins = False, False
            if not vessel_insight.empty:
                vessel_insight.to_excel(writer, sheet_name='KPI_vessel_use', index=True)
                v_ins = True
            if not cost_insight.empty:
                cost_insight.to_excel(writer, sheet_name='KPI_failure_contribution', index=True)
                c_ins = True
            
            if v_ins or c_ins:
                worksheet = writer.sheets['KPI_failure_contribution']
            if c_ins:  
                # Write notes
                worksheet.cell(row=len(cost_insight)+3, column=1, value="The failure contribution is made evaluating for each failure the FR*devices")
                worksheet.cell(row=len(cost_insight)+4, column=1, value="This show the failure per year that occur in the farm and they are correlated with the vessel_type")
                worksheet.cell(row=len(cost_insight)+4, column=1, value="All the values must be multiplied for the number found only once in the vessels type")