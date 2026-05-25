# test_return_statistics_runs.py

import os
import unittest
import tempfile
from types import SimpleNamespace
from datetime import datetime

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from unittest.mock import patch, MagicMock

from oriom.core.statistical_analysis.final_run_statistics import (
    return_statistics_runs,
)


class DummyResults(SimpleNamespace):
    """Simple container for the attributes used by return_statistics_runs."""
    pass


class TestReturnStatisticsRuns(unittest.TestCase):
    def setUp(self):
        self.tmp_ctx = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp_ctx.cleanup)
        self.save_dir = self.tmp_ctx.name

        # Common parameters
        self.n_lifetime = 2
        self.n_runs = 1

        # Simple electricity cost dict
        self.electricity_cost_dict = {
            "wt": 100.0,  # €/MWh
            "pv": 50.0,
            "wec": 80.0,
        }

        # Simple vessels / operations_total placeholders
        self.vessels = []
        self.operations_total = []
        self.find_element_class = {}  # passed to yearly_vessel_cost but mocked

    def _build_dummy_tot_cost_df(self):
        """
        Build a minimal lifetime cost DataFrame as expected by dfs_tot_cost_list:
        - multiple rows with vessel_id and lifetime_direct_costs
        - last row with aggregated totals (tot_* columns).
        """
        data = [
            # vessel_id, lifetime_direct_costs,   tot_* only used on last row
            ("insurance", 1000.0, np.nan, np.nan, np.nan, np.nan, np.nan),
            ("port",      2000.0, np.nan, np.nan, np.nan, np.nan, np.nan),
            ("technician",3000.0, np.nan, np.nan, np.nan, np.nan, np.nan),
            (
                "TOTAL",
                6000.0,
                700.0,   # tot_technicians_costs
                800.0,   # tot_vessel_costs
                900.0,   # tot_mobilization_costs
                400.0,   # tot_rov_costs
                500.0,   # tot_part_costs
            ),
        ]
        return pd.DataFrame(
            data,
            columns=[
                "vessel_id",
                "lifetime_direct_costs",
                "tot_technicians_costs",
                "tot_vessel_costs",
                "tot_mobilization_costs",
                "tot_rov_costs",
                "tot_part_costs",
            ],
        )

    def _build_dummy_yearly_cost_df_multiindex(self):
        """
        Build a yearly cost DataFrame with MultiIndex columns:
        - first column 'vessel_id'
        - other columns like (year, 'direct_costs') and (year, 'n_days').
        """
        vessel_ids = ["vA"]
        cols = pd.MultiIndex.from_tuples(
            [
                (2020, "direct_costs"),
                (2020, "n_days"),
                (2021, "direct_costs"),
                (2021, "n_days"),
            ],
            names=["year", "metric"],
        )
        data = np.array([[1000.0, 10.0, 1100.0, 11.0]])  # shape (1,4)
        df = pd.DataFrame(data, columns=cols)
        df.insert(0, "vessel_id", vessel_ids)
        return df

    def _build_dummy_yearly_cost_df_flat(self):
        """
        Build a yearly cost DataFrame in 'flat' format (for recycled=True):
        columns like 'vessel_id', '2020', '2020.1' etc.
        """
        return pd.DataFrame(
            {
                "vessel_id": ["vA"],
                "2020": [1000.0],    # direct_costs
                "2020.1": [10.0],    # n_days
                "2021": [1100.0],    # direct_costs
                "2021.1": [11.0],    # n_days
            }
        )

    def _build_dummy_ctv_df_list(self):
        """Create a simple list of DataFrames for dfs_ctv_list."""
        df_ctv = pd.DataFrame(
            {
                "year": [2020, 2021],
                "cost": [100.0, 120.0],
            }
        )
        return [df_ctv]

    def _build_dummy_type_cost_list(self):
        """Create a simple list of DataFrames for kpi_om_type_cost_list."""
        df = pd.DataFrame(
            {
                "description": ["inspection", "correction"],
                "values": [1.0e6, 2.0e6],
            }
        )
        return [df]

    def _build_dummy_energy_yearly_dict(self):
        """
        Build dfs_energy_yearly_dict:
        keys -> list of yearly DataFrames (one per simulation).
        Only 'wind' is needed to exercise the path.
        """
        df_wind = pd.DataFrame(
            {
                "En_max_kWh": [1000.0, 1100.0],
                "En_loss_kWh": [100.0, 120.0],
                "En_availability": [0.9, 0.91],
                "Time_availability": [0.95, 0.96],
            }
        )
        return {
            "farm_availability_wind": [df_wind]
        }

    def _build_dummy_energy_yearly_month_dict(self):
        """
        Build dfs_energy_yearly_month_dict:
        keys -> list of monthly DataFrames (one per simulation).
        Only one key is enough to exercise the path.
        """
        df_month = pd.DataFrame(
            {
                "month": [1, 2],
                "En_max_kWh": [100.0, 110.0],
                "En_loss_kWh": [10.0, 12.0],
                "En_availability": [0.9, 0.91],
                "Time_availability": [0.95, 0.96],
            }
        )
        return {
            "energy_yield_yearly_wind": [df_month]
        }

    def _build_dummy_kpi_insight(self):
        """
        Build cost_insight and vessel_insight DataFrames for KPI_Insight.kpi_insight.
        """
        cost_insight = pd.DataFrame(
            {
                "failure": ["F1", "F2"],
                "contribution": [0.6, 0.4],
            }
        ).set_index("failure")

        vessel_insight = pd.DataFrame(
            {
                "reuse %": [0.3],
                "merge %": [0.2],
                "yearly day effective": [50.0],
            },
            index=["ctv"],
        )

        return cost_insight, vessel_insight

    def test_return_statistics_runs_happy_path(self):
        """
        Full happy-path test:
        - multiindex yearly costs
        - CTV, type costs, yearly + monthly energy
        - KPI_Insight is called and Excel files are produced.
        """
        df_tot_cost = self._build_dummy_tot_cost_df()
        df_year_multi = self._build_dummy_yearly_cost_df_multiindex()
        dfs_ctv_list = self._build_dummy_ctv_df_list()
        kpi_type_cost_list = self._build_dummy_type_cost_list()
        dfs_energy_yearly_dict = self._build_dummy_energy_yearly_dict()
        dfs_energy_yearly_month_dict = self._build_dummy_energy_yearly_month_dict()
        cost_insight, vessel_insight = self._build_dummy_kpi_insight()

        results = DummyResults(
            dfs_tot_cost_list=[df_tot_cost],
            dfs_tot_yearly_cost_list=[df_year_multi],
            dfs_ctv_list=dfs_ctv_list,
            kpi_om_type_cost_list=kpi_type_cost_list,
            dfs_energy_yearly_dict=dfs_energy_yearly_dict,
            dfs_energy_yearly_month_dict=dfs_energy_yearly_month_dict,
            dfs_vessel_fuel_usage={}
        )

        fuel_add = {"vA": 10.0}
        mobilisation_add = {"vA": 20.0}

        # Patch external side-effects: plots, CSVs, KPI_Insight
        module_path = "oriom.core.statistical_analysis.final_run_statistics"
        with patch(f"{module_path}.aux_functions.save_file_csv") as m_save_csv, \
             patch(f"{module_path}.final_economic_graphs") as m_econ_graphs, \
             patch(f"{module_path}.report_graphs") as m_report_graphs, \
             patch(f"{module_path}.KPI_Insight") as m_kpi_cls:

            # Configure KPI_Insight mock
            kpi_instance = MagicMock()
            kpi_instance.kpi_insight.return_value = (cost_insight, vessel_insight)
            m_kpi_cls.return_value = kpi_instance

            # Call function under test
            return_statistics_runs(
                n_lifetime=self.n_lifetime,
                find_element_class=self.find_element_class,
                results_dict=results,
                fuel_add=fuel_add,
                mobilisation_add=mobilisation_add,
                electricity_cost_dict=self.electricity_cost_dict,
                n_runs=self.n_runs,
                vessels=self.vessels,
                operations_total=self.operations_total,
                recycled=False,
                save_dir=self.save_dir,
            )

            # Check main Excel output is created
            avg_path = os.path.join(self.save_dir, "Average_results.xlsx")
            self.assertTrue(os.path.isfile(avg_path))

            try:
                from oriom.core.functions.private.KPI_Insight import KPI_Insight
                # Check KPI insight Excel is created
                kpi_path = os.path.join(self.save_dir, "KPI_insight.xlsx")
                self.assertTrue(os.path.isfile(kpi_path))
                # KPI_Insight should be instantiated and called
                m_kpi_cls.assert_called_once_with(N_SIMULATION=self.n_runs, n_lifetime=self.n_lifetime)
                kpi_instance.kpi_insight.assert_called_once()
            except (ImportError, ModuleNotFoundError):
                pass

            # Open Average_results and check relevant sheets and values
            wb = load_workbook(avg_path)
            sheet_names = wb.sheetnames

            self.assertIn("Lifetime_results", sheet_names)
            self.assertIn("Lifetime_costs", sheet_names)
            self.assertIn("Yearly_costs", sheet_names)
            self.assertIn("CTV_yearly_strategy", sheet_names)
            self.assertIn("Energy_results", sheet_names)

            # Check that lifetime_direct_cost row exists and value matches expectation
            ws_costs = wb["Lifetime_costs"]
            lifetime_direct_cost_value = None
            for row in ws_costs.iter_rows(min_row=2, values_only=True):
                if row[0] == "lifetime_direct_cost":
                    lifetime_direct_cost_value = row[1]
                    break
            self.assertIsNotNone(lifetime_direct_cost_value)

            # Manually compute expected lifetime_direct_cost:
            # from _build_dummy_tot_cost_df():
            #   lifetime_vessels_cost = 800 + 2*10 = 820
            #   lifetime_mobilisation_cost = 900 + 2*20 = 940
            #   lifetime_rov_cost = 400
            #   lifetime_technician_cost = 700
            #   lifetime_repair_cost = 500
            #   lifetime_fixed_port_cost = 2000
            #   lifetime_fixed_tech_cost = 3000
            #   lifetime_fixed_insurance_cost = 1000
            expected = 820 + 940 + 400 + 700 + 500 + 2000 + 3000 + 1000  # 9360
            self.assertAlmostEqual(lifetime_direct_cost_value, expected, places=6)

            # Some of the plotting helpers should have been called at least once
            self.assertTrue(m_econ_graphs.lifetime_cost.called)
            self.assertTrue(m_econ_graphs.yearly_vessel_cost.called)
            self.assertTrue(m_report_graphs.farm_availability.called)
            self.assertTrue(m_report_graphs.energy_yield.called)

    def test_return_statistics_runs_recycled_yearly_structure(self):
        """
        Test the `recycled=True` branch: df.yearly is in flat format and must
        be converted by `restructure_df_year` without raising errors.
        """
        df_tot_cost = self._build_dummy_tot_cost_df()
        df_year_flat = self._build_dummy_yearly_cost_df_flat()

        results = DummyResults(
            dfs_tot_cost_list=[df_tot_cost],
            dfs_tot_yearly_cost_list=[df_year_flat],
            dfs_ctv_list=[],
            kpi_om_type_cost_list=[],
            dfs_energy_yearly_dict={},
            dfs_energy_yearly_month_dict={},
            dfs_vessel_fuel_usage={}
        )

        fuel_add = {"vA": 5.0}
        mobilisation_add = {"vA": 5.0}

        module_path = "oriom.core.statistical_analysis.final_run_statistics"
        with patch(f"{module_path}.aux_functions.save_file_csv") as m_save_csv, \
             patch(f"{module_path}.final_economic_graphs") as m_econ_graphs, \
             patch(f"{module_path}.report_graphs") as m_report_graphs, \
             patch(f"{module_path}.KPI_Insight") as m_kpi_cls:

            kpi_instance = MagicMock()
            # Very simple empty DataFrames
            kpi_instance.kpi_insight.return_value = (
                pd.DataFrame(index=[], columns=[]),
                pd.DataFrame(index=[], columns=[]),
            )
            m_kpi_cls.return_value = kpi_instance

            # Call function under test
            return_statistics_runs(
                n_lifetime=self.n_lifetime,
                find_element_class=self.find_element_class,
                results_dict=results,
                fuel_add=fuel_add,
                mobilisation_add=mobilisation_add,
                electricity_cost_dict=self.electricity_cost_dict,
                n_runs=self.n_runs,
                vessels=self.vessels,
                operations_total=self.operations_total,
                recycled=True,  # <--- triggers restructure_df_year
                save_dir=self.save_dir,
            )

            # Check that Average_results.xlsx exists and has the Yearly_costs sheet
            avg_path = os.path.join(self.save_dir, "Average_results.xlsx")
            self.assertTrue(os.path.isfile(avg_path))
            wb = load_workbook(avg_path)
            self.assertIn("Yearly_costs", wb.sheetnames)

            # No energy dict -> Energy_results sheet may not exist
            # but Lifetime_results and Lifetime_costs must be present
            self.assertIn("Lifetime_results", wb.sheetnames)
            self.assertIn("Lifetime_costs", wb.sheetnames)


if __name__ == "__main__":
    unittest.main(verbosity=2)
